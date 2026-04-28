#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V55 — LAS CUATRO LEYES DE LA IA HUMANA
Dr. Óscar Gallego Castilla — Febrero 2026
═══════════════════════════════════════════
MODELO MATEMÁTICO COMPLETO v2 — Parte VI (Capítulos 30-35)
13 pestañas: GUIA, Cap30_Marco, Cap31_Teoria, Cap31_Ejemplos,
Cap32_Teoria, Cap32_Ejemplos, Cap33_Teoria, Cap33_Ejemplos,
Cap34_Teoria, Cap34_Ejemplos, Cap35_IHG, Cap35_Validacion, Dashboard

TODOS los datos numéricos proceden EXCLUSIVAMENTE del texto del V55.
Ningún dato es inventado. Cada celda referencia su sección/tabla.

Requisitos:  pip install openpyxl
Ejecución:   python v55_modelo_completo_v2.py
Salida:      V55_Modelo_Completo_v2.xlsx (13 pestañas)

Google Colab:
  !pip install openpyxl
  # Ejecutar este script
  from google.colab import files
  files.download("V55_Modelo_Completo_v2.xlsx")
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import math

wb = Workbook()

# ═══════════════════════════════════════════════════════════════
# ESTILOS GLOBALES
# ═══════════════════════════════════════════════════════════════
TF = Font(name='Calibri', size=13, bold=True, color='1B2631')
HF = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
SF = Font(name='Calibri', size=10, bold=True)
NF = Font(name='Calibri', size=10)
FF = Font(name='Consolas', size=10)
RF = Font(name='Calibri', size=12, bold=True, color='006100')
SM = Font(name='Calibri', size=9, italic=True, color='566573')
HFL = PatternFill('solid', fgColor='2C3E50')
INP = PatternFill('solid', fgColor='FFF9C4')
CAL = PatternFill('solid', fgColor='BBDEFB')
RES = PatternFill('solid', fgColor='C8E6C9')
WRN = PatternFill('solid', fgColor='FFCDD2')
ORG = PatternFill('solid', fgColor='FFE0B2')
YEL = PatternFill('solid', fgColor='FFF9C4')
GRN = PatternFill('solid', fgColor='C8E6C9')
GRN2 = PatternFill('solid', fgColor='A5D6A7')
WHT = PatternFill('solid', fgColor='F5F5F5')
LIL = PatternFill('solid', fgColor='E1BEE7')
TB = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CT = Alignment(horizontal='center', vertical='center', wrap_text=True)
LW = Alignment(horizontal='left', vertical='center', wrap_text=True)

def sc(ws, r, c, v, font=NF, fill=None, al=CT, brd=TB, fmt=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = font
    if fill: cl.fill = fill
    cl.alignment = al
    if brd: cl.border = brd
    if fmt: cl.number_format = fmt
    return cl

def hdr(ws, r, cols):
    for c, v in enumerate(cols, 1):
        sc(ws, r, c, v, font=HF, fill=HFL)

def ttl(ws, r, c, txt, me=None):
    sc(ws, r, c, txt, font=TF, fill=None, brd=None, al=LW)
    if me: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=me)

def note(ws, r, c, txt, me=None):
    sc(ws, r, c, txt, font=SM, fill=None, brd=None, al=LW)
    if me: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=me)

def setw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ████████████████████████████████████████████████████████████████
# PESTAÑA 1: GUÍA DIDÁCTICA
# ████████████████████████████████████████████████████████████████
ws0 = wb.active
ws0.title = "GUIA"
ws0.sheet_properties.tabColor = "2C3E50"
setw(ws0, [4, 22, 42, 42])
r = 1
ttl(ws0, r, 1, "V55 — MODELO MATEMÁTICO COMPLETO v2", 4); r += 1
note(ws0, r, 1, "Dr. Óscar Gallego Castilla — Las Cuatro Leyes de la IA Humana — Parte VI (Caps 30-35)", 4); r += 2

ttl(ws0, r, 1, "CÓDIGO DE COLORES", 4); r += 1
hdr(ws0, r, ["", "Color", "Significado", "Instrucción"]); r += 1
for cod, sig, inst, fl in [
    ("🟡 AMARILLO", "INPUTS EDITABLES — datos del sistema a evaluar", "Modifique para evaluar SU sistema", INP),
    ("⬜ GRIS", "PESOS Y PARÁMETROS propuestos por el autor (V55)", "Puede ajustar, documente desviaciones", WHT),
    ("🔵 AZUL", "FÓRMULAS CALCULADAS — operaciones intermedias", "No editar: contienen fórmulas Excel", CAL),
    ("🟢 VERDE", "RESULTADOS FINALES — métricas de salida", "Valor final + dictamen de cumplimiento", RES),
    ("🟣 LILA", "TABLAS DE REFERENCIA — escalas, umbrales, rangos", "Consultar para interpretar resultados", LIL),
    ("🔴 ROJO", "ALERTAS — zona crítica o incumplimiento", "Requiere intervención urgente", WRN),
]:
    sc(ws0, r, 2, cod, fill=fl); sc(ws0, r, 3, sig); sc(ws0, r, 4, inst); r += 1
r += 1

ttl(ws0, r, 1, "MAPA DE PESTAÑAS", 4); r += 1
hdr(ws0, r, ["#", "Pestaña", "Contenido", "Tablas / Secciones V55"]); r += 1
tabs = [
    ("1", "GUIA", "Esta guía didáctica", "—"),
    ("2", "Cap30_Marco", "Arquitectura dos niveles, D1-D4, EUC, Declaración Residuo", "T44, T45, T46, T47 — §7.1"),
    ("3", "Cap31_Teoria", "MUI, fórmulas I, ρₕ, EUC personalización", "T48, §7.2.3-4"),
    ("4", "Cap31_Ejemplos", "María+Interestelar, Tabla 49, ρₕ, T50 convergencia", "T49, T50 — §7.2.5-7"),
    ("5", "Cap32_Teoria", "Fórmulas α, M(d), R, AHP moral, T52-56", "T52-56, T60 — §7.3"),
    ("6", "Cap32_Ejemplos", "Préstamo, moderación, diagnóstico, erosión", "T57-59, T61 — §7.3.6-17"),
    ("7", "Cap33_Teoria", "Fórmulas CEE, ICD, IDP, ICT, AHP transparencia", "T64-73 — §7.4"),
    ("8", "Cap33_Ejemplos", "SHAP préstamo, ICD 3 plataformas, IDP, ICT", "T64-74 — §7.4.4-15"),
    ("9", "Cap34_Teoria", "Clasificador moral, sigmoide, φ, pesos w, umbral θ", "T76-77, T81 — §7.5"),
    ("10", "Cap34_Ejemplos", "3 decisiones φ, 9 decisiones espectro, crédito, Tasa A", "T78-83, T85 — §7.5.7-26"),
    ("11", "Cap35_IHG", "Fórmula IHG, AHP pesos leyes, condiciones necesarias", "T89-92 — §7.6.4-9"),
    ("12", "Cap35_Validacion", "Patologías, sensibilidad pesos, erosión temporal", "T93-94 — §7.6.10-14"),
    ("13", "Dashboard", "Panel resumen final + dictamen certificación", "Integra todo"),
]
for t in tabs:
    for c, v in enumerate(t, 1):
        sc(ws0, r, c, v)
    r += 1
r += 1

ttl(ws0, r, 1, "FLUJO DE CERTIFICACIÓN (§7.1.2)", 4); r += 1
for paso in [
    "Certificación = (D1 ∧ D2 ∧ D3 ∧ D4) ∧ (IHG ≥ 0.60) ∧ (C1 ∧ C2 ∧ C3)",
    "PASO 1: Verificar D1-D4 (pestaña Cap30_Marco) → si alguno falla, DESCALIFICADO",
    "PASO 2: Calcular I, R, ICT, A (pestañas Cap31-34) → cada uno con su EUC",
    "PASO 3: Calcular IHG + 3 condiciones (pestaña Cap35_IHG)",
    "PASO 4: Panel final (pestaña Dashboard) → DICTAMEN",
]:
    note(ws0, r, 1, paso, 4); r += 1
r += 1
ttl(ws0, r, 1, "NOTA IMPORTANTE", 4); r += 1
note(ws0, r, 1, "Todos los datos numéricos precargados son los EJEMPLOS ILUSTRATIVOS del V55.", 4); r += 1
note(ws0, r, 1, "Para auditar un sistema real, sustituya SOLO las celdas amarillas con datos del sistema.", 4); r += 1
note(ws0, r, 1, "Todas las formulaciones son especulativas y requieren validación empírica (§7.1.1).", 4)


# ████████████████████████████████████████████████████████████████
# PESTAÑA 2: CAP 30 — MARCO FUNDAMENTOS
# ████████████████████████████████████████████████████████████████
ws1 = wb.create_sheet("Cap30_Marco")
ws1.sheet_properties.tabColor = "795548"
setw(ws1, [4, 32, 35, 32, 28])
r = 1
ttl(ws1, r, 1, "CAPÍTULO 30 — FUNDAMENTOS MATEMÁTICOS (§7.1)", 5); r += 2

# Tabla 44: Condiciones deontológicas
ttl(ws1, r, 1, "NIVEL 1: CONDICIONES DEONTOLÓGICAS (Tabla 44, §7.1.3)", 5); r += 1
hdr(ws1, r, ["", "Condición", "Ley que protege", "Pregunta de verificación", "Método de auditoría"]); r += 1
for cond, ley, preg, met in [
    ("D1: No instrumentalización", "1ª (Persona)", "¿Función objetivo incluye bienestar del usuario?", "Auditoría de función objetivo"),
    ("D2: Override humano", "2ª (Aumento)", "¿Existe mecanismo de reversión humana operativo?", "Verificación técnica de flujo"),
    ("D3: Explicación accesible", "3ª (Transparencia)", "¿Toda decisión significativa tiene explicación?", "Test comprensión con usuarios"),
    ("D4: Abstención ante lo moral", "4ª (Límite)", "¿Existe mecanismo de escalamiento a humano?", "Verificación técnica + auditoría"),
]:
    sc(ws1, r, 2, cond); sc(ws1, r, 3, ley, fill=LIL); sc(ws1, r, 4, preg); sc(ws1, r, 5, met); r += 1
r += 1

# Verificación D1-D4
ttl(ws1, r, 1, "VERIFICACIÓN D1-D4 (celdas amarillas: 1=Sí, 0=No)", 5); r += 1
hdr(ws1, r, ["", "Condición", "¿Se cumple?", "Resultado", ""]); r += 1
d_rows = []
for d in ["D1: No instrumentalización", "D2: Override humano operativo",
          "D3: Explicación accesible", "D4: Abstención ante lo moral"]:
    sc(ws1, r, 2, d); sc(ws1, r, 3, 1, fill=INP)
    sc(ws1, r, 4, None, fill=RES)
    ws1.cell(r, 4).value = f'=IF(C{r}=1,"✓ CUMPLE","✗ FALLA → DESCALIFICADO")'
    ws1.cell(r, 4).font = RF; d_rows.append(r); r += 1
sc(ws1, r, 2, "NIVEL 1 GLOBAL", font=SF); sc(ws1, r, 3, None, fill=RES)
cond_and = " * ".join([f"C{dr}" for dr in d_rows])
ws1.cell(r, 3).value = f'=IF({cond_and}=1,"✓ TODAS CUMPLEN → PROCEDER NIVEL 2","✗ DESCALIFICADO")'
ws1.cell(r, 3).font = RF; rN1 = r; r += 2

# Tabla 45: EUC (§7.1.4)
ttl(ws1, r, 1, "ESCALA UNIVERSAL DE CUMPLIMIENTO — EUC (Tabla 45, §7.1.4)", 5); r += 1
hdr(ws1, r, ["", "Nivel", "Rango", "Significado cualitativo", "Acción requerida"]); r += 1
for niv, rng, sig, acc, fl in [
    ("Crítico", "[0.00, 0.25)", "Violación grave y sistémica", "Suspensión hasta corrección", WRN),
    ("Insuficiente", "[0.25, 0.50)", "Violación en aspectos sustanciales", "Intervención urgente", ORG),
    ("Aceptable", "[0.50, 0.70)", "Cumplimiento mínimo con riesgos", "Intervención a medio plazo", YEL),
    ("Bueno", "[0.70, 0.85)", "Cumplimiento con margen de mejora", "Optimización programada", GRN),
    ("Excelente", "[0.85, 1.00]", "Cumplimiento robusto y consistente", "Mantenimiento + mejora continua", GRN2),
]:
    sc(ws1, r, 2, niv, fill=fl); sc(ws1, r, 3, rng, fill=fl)
    sc(ws1, r, 4, sig, fill=fl); sc(ws1, r, 5, acc, fill=fl); r += 1
r += 1

# Tabla 46: Aplicación EUC por ley
ttl(ws1, r, 1, "APLICACIÓN DE LA EUC A LAS MÉTRICAS POR LEY (Tabla 46, §7.1.5)", 5); r += 1
hdr(ws1, r, ["", "Ley", "Métrica", "Umbral mínimo", "Nivel EUC requerido"]); r += 1
for ley, met, umb, niv in [
    ("1ª Persona", "Índice Personalización (I)", "I ≥ 0.70", "Bueno o superior"),
    ("2ª Aumento", "Ratio Decisión Humana (R)", "R ≥ 0.70", "Bueno o superior"),
    ("3ª Transparencia", "Índice Compuesto Transp. (ICT)", "ICT ≥ 0.75", "Bueno (zona alta)"),
    ("4ª Límite", "Tasa Abstención Correcta (A)", "A ≥ 0.95", "Excelente (zona alta)"),
]:
    sc(ws1, r, 2, ley, fill=LIL); sc(ws1, r, 3, met); sc(ws1, r, 4, umb); sc(ws1, r, 5, niv); r += 1
r += 1

# Tabla 47: Resumen Cap 30
ttl(ws1, r, 1, "RESUMEN CAPÍTULO 30 (Tabla 47, §7.1.8)", 5); r += 1
hdr(ws1, r, ["", "Concepto", "Descripción", "Aplicación", "Riesgo si se ignora"]); r += 1
for con, desc, apl, rie in [
    ("Arquitectura dos niveles", "Nivel 1 (deontológico) + Nivel 2 (cuantitativo)", "Respeta incondicionalidad dignidad + gradualidad implementación", "Tratar dignidad como variable continua"),
    ("Condiciones D1-D4", "Verificaciones binarias previas", "Filtro categórico que ninguna puntuación elude", "Compensar violaciones con excelencia en otras dimensiones"),
    ("EUC", "Cinco niveles en [0,1]", "Comparabilidad entre sistemas y disciplinas", "Escalas ad hoc incomparables"),
    ("Declaración de Residuo", "Documentación de lo que métricas no capturan", "Salvaguarda contra confundir mapa con territorio", "Creer que IHG mide dignidad cuando mide proxies"),
    ("Principios metodológicos", "Incertidumbre, validez, coherencia, monitoreo, adaptabilidad", "Marco de rigor para capítulos siguientes", "Propuestas sin criterio de evaluación propia"),
]:
    sc(ws1, r, 2, con); sc(ws1, r, 3, desc); sc(ws1, r, 4, apl); sc(ws1, r, 5, rie); r += 1


# ████████████████████████████████████████████████████████████████
# PESTAÑA 3: CAP 31 — TEORÍA LEY DE LA PERSONA
# ████████████████████████████████████████████████████████████████
ws2 = wb.create_sheet("Cap31_Teoria")
ws2.sheet_properties.tabColor = "1F4E79"
setw(ws2, [4, 38, 35, 28, 20, 18])
r = 1
ttl(ws2, r, 1, "CAPÍTULO 31 — MÉTRICAS PARA LA LEY DE LA PERSONA (§7.2)", 6); r += 2

# MUI (Tabla 48)
ttl(ws2, r, 1, "1. MODELO DE USUARIO INDIVIDUAL — MUI (Tabla 48, §7.2.3)", 6); r += 1
hdr(ws2, r, ["#", "Componente", "Fórmula", "Significado", "Ref V55", ""]); r += 1
for i, (c, f, s, ref) in enumerate([
    ("Predicción total", "f(x) = g(x) + h(x, Mᵤ)", "Salida completa del sistema", "§7.2.3"),
    ("Componente poblacional", "g(x)", "Predicción sin conocer al usuario", "§7.2.3"),
    ("Componente personal", "h(x, Mᵤ)", "Ajuste por memoria episódica", "§7.2.3"),
    ("Memoria episódica", "Mᵤ = {Pᵤ, Hᵤ, Cᵤ}", "Preferencias + Historia + Correcciones", "§7.2.3"),
], 1):
    sc(ws2, r, 1, i); sc(ws2, r, 2, c); sc(ws2, r, 3, f, font=FF); sc(ws2, r, 4, s); sc(ws2, r, 5, ref); r += 1
r += 1

# Métricas (§7.2.4)
ttl(ws2, r, 1, "2. MÉTRICAS (§7.2.4)", 6); r += 1
hdr(ws2, r, ["#", "Métrica", "Fórmula", "Umbral", "Ref V55", ""]); r += 1
for i, (m, f, u, ref) in enumerate([
    ("Índice de Personalización", "I = Var(h) / Var(f)", "I ≥ 0.70", "§7.2.4"),
    ("Coeficiente adecuación", "ρₕ = Cov(h,s) / (σₕ × σₛ)", "ρₕ ≥ 0.30", "§7.2.4.1"),
    ("Piso varianza absoluta", "Var(h) ≥ ε_min", "ε_min = 0.005", "§7.2.4.2"),
], 1):
    sc(ws2, r, 1, i); sc(ws2, r, 2, m); sc(ws2, r, 3, f, font=FF); sc(ws2, r, 4, u); sc(ws2, r, 5, ref); r += 1
r += 1

# EUC para I
ttl(ws2, r, 1, "3. EUC APLICADA AL ÍNDICE I", 6); r += 1
hdr(ws2, r, ["#", "Zona", "Rango I", "Interpretación", "", ""]); r += 1
for i, (z, rng, interp, fl) in enumerate([
    ("Insuficiente", "I < 0.30", "Personalización mínima o nula", WRN),
    ("Básico", "0.30 ≤ I < 0.50", "Incipiente pero insuficiente", ORG),
    ("Aceptable", "0.50 ≤ I < 0.70", "Parcial, margen significativo", YEL),
    ("Bueno", "0.70 ≤ I < 0.85", "✓ Genuina (umbral cumplido)", GRN),
    ("Excelente", "I ≥ 0.85", "Avanzada, referencia del sector", GRN2),
], 1):
    sc(ws2, r, 1, i, fill=fl); sc(ws2, r, 2, z, fill=fl); sc(ws2, r, 3, rng, fill=fl); sc(ws2, r, 4, interp, fill=fl); r += 1
r += 1

# Convergencia temporal — Tabla 50 (§7.2.7)
ttl(ws2, r, 1, "4. CONVERGENCIA TEMPORAL I(t) — Tabla 50 (§7.2.7)", 6); r += 1
note(ws2, r, 1, "I(t) → I ≥ 0.70 cuando t → T_dominio", 6); r += 1
hdr(ws2, r, ["#", "Dominio", "T_dominio estimado", "Justificación", "", ""]); r += 1
for i, (dom, t, just) in enumerate([
    ("Recomendación contenido", "20-50 interacciones", "Alta frecuencia, preferencias estables"),
    ("Asistente IA conversacional", "10-30 conversaciones", "Interacciones ricas, aprendizaje rápido"),
    ("E-commerce", "30-80 interacciones", "Preferencias diversas por categoría"),
    ("Educación personalizada", "15-40 sesiones", "Necesidades emergen progresivamente"),
    ("Salud y bienestar", "40-100 interacciones", "Dominio sensible, más cautela"),
], 1):
    sc(ws2, r, 1, i); sc(ws2, r, 2, dom, fill=LIL); sc(ws2, r, 3, t); sc(ws2, r, 4, just); r += 1


# ████████████████████████████████████████████████████████████████
# PESTAÑA 4: CAP 31 — EJEMPLOS NUMÉRICOS
# ████████████████████████████████████████████████████████████████
ws3 = wb.create_sheet("Cap31_Ejemplos")
ws3.sheet_properties.tabColor = "2E86C1"
setw(ws3, [4, 20, 14, 14, 14, 4, 18, 18, 4, 30])
r = 1

# --- Ejemplo María + Interestelar (§7.2.5) ---
ttl(ws3, r, 1, "EJEMPLO 7 — María + Interestelar (§7.2.5)", 10); r += 2
ttl(ws3, r, 1, "PASO 1: g(x) — Modelo base", 5); r += 1
sc(ws3, r, 2, 'g("Interestelar")'); sc(ws3, r, 3, 72, fill=INP, fmt='0.00')
sc(ws3, r, 4, "pts"); sc(ws3, r, 5, "Usuario promedio"); rg = r; r += 2

ttl(ws3, r, 1, "PASO 2: h(x, Mᵤ) — Ajuste individual", 5); r += 1
sc(ws3, r, 2, "Pᵤ: Pref. explícita"); sc(ws3, r, 3, 15, fill=INP, fmt='0.00')
sc(ws3, r, 10, '"Me encanta la ciencia ficción" → +15'); rp = r; r += 1
sc(ws3, r, 2, "Hᵤ: Historia"); sc(ws3, r, 3, 8, fill=INP, fmt='0.00')
sc(ws3, r, 10, "3 películas de Nolan completas → +8"); rh = r; r += 1
sc(ws3, r, 2, "Cᵤ: Corrección"); sc(ws3, r, 3, -5, fill=INP, fmt='0.00')
sc(ws3, r, 10, "Mala recomendación película >150min → −5"); rc = r; r += 1
sc(ws3, r, 2, "h total", font=SF); sc(ws3, r, 3, None, fill=CAL, fmt='0.00')
ws3.cell(r, 3).value = f"=C{rp}+C{rh}+C{rc}"; rht = r; r += 2

ttl(ws3, r, 1, "PASO 3: f(x) = g(x) + h(x,Mᵤ)", 5); r += 1
sc(ws3, r, 2, "f(Interestelar)", font=SF); sc(ws3, r, 3, None, fill=RES, fmt='0.00')
ws3.cell(r, 3).value = f"=C{rg}+C{rht}"; ws3.cell(r, 3).font = RF
sc(ws3, r, 10, "V55: 72 + 18 = 90 puntos (§7.2.5)"); r += 3

# --- Tabla 49: Cálculo Índice I (§7.2.6) ---
ttl(ws3, r, 1, "EJEMPLO 8 — Índice I paso a paso (Tabla 49, §7.2.6)", 10); r += 1
note(ws3, r, 1, "10 predicciones de sistema de contenido educativo. Datos EXACTOS del V55 Tabla 49.", 10); r += 1
hdr(ws3, r, ["", "Predicción", "f(x)", "g(x)", "h(x,Mᵤ)", "", "f−f̄", "h−h̄", "", "Ref"]); r += 1

# Datos EXACTOS de la Tabla 49 del V55
t49 = [
    (1, 0.82, 0.55, 0.27), (2, 0.91, 0.60, 0.31), (3, 0.45, 0.50, -0.05),
    (4, 0.78, 0.58, 0.20), (5, 0.93, 0.52, 0.41), (6, 0.30, 0.48, -0.18),
    (7, 0.87, 0.55, 0.32), (8, 0.65, 0.57, 0.08), (9, 0.88, 0.54, 0.34),
    (10, 0.72, 0.53, 0.19),
]
data_start = r
for pred, fx, gx, hx in t49:
    sc(ws3, r, 2, pred); sc(ws3, r, 3, fx, fill=INP, fmt='0.0000')
    sc(ws3, r, 4, gx, fill=INP, fmt='0.0000'); sc(ws3, r, 5, hx, fill=INP, fmt='0.0000')
    sc(ws3, r, 7, None, fill=CAL, fmt='0.0000')
    sc(ws3, r, 8, None, fill=CAL, fmt='0.0000')
    r += 1
data_end = r - 1; r += 1

# Medias (V55: f̄ = 0.731, h̄ = 0.189)
sc(ws3, r, 2, "Media", font=SF)
sc(ws3, r, 3, None, fill=CAL, fmt='0.0000')
ws3.cell(r, 3).value = f"=AVERAGE(C{data_start}:C{data_end})"
sc(ws3, r, 5, None, fill=CAL, fmt='0.0000')
ws3.cell(r, 5).value = f"=AVERAGE(E{data_start}:E{data_end})"
sc(ws3, r, 10, "V55: f̄=0.731, h̄=0.189")
rmf = r; rmh = r; r += 1

# Rellenar desviaciones
for rr in range(data_start, data_end + 1):
    ws3.cell(rr, 7).value = f"=C{rr}-C${rmf}"
    ws3.cell(rr, 8).value = f"=E{rr}-E${rmh}"

# Varianzas (V55: Var(f)=0.0395, Var(h)=0.0356)
sc(ws3, r, 2, "Varianza", font=SF)
sc(ws3, r, 3, None, fill=CAL, fmt='0.0000')
ws3.cell(r, 3).value = f"=VAR.P(C{data_start}:C{data_end})"
sc(ws3, r, 5, None, fill=CAL, fmt='0.0000')
ws3.cell(r, 5).value = f"=VAR.P(E{data_start}:E{data_end})"
sc(ws3, r, 10, "V55: Var(f)=0.0395, Var(h)=0.0356")
rvf = r; rvh = r; r += 2

# I = Var(h) / Var(f)
sc(ws3, r, 2, "I = Var(h)/Var(f)", font=SF)
sc(ws3, r, 3, None, fill=RES, fmt='0.0000')
ws3.cell(r, 3).value = f"=E{rvh}/C{rvf}"
ws3.cell(r, 3).font = RF
sc(ws3, r, 5, "Umbral ≥ 0.70")
sc(ws3, r, 7, None, fill=RES)
ws3.cell(r, 7).value = f'=IF(C{r}>=0.70,"✓ CUMPLE","✗ NO CUMPLE")'
ws3.cell(r, 7).font = RF
sc(ws3, r, 10, "V55: I = 0.0356/0.0395 ≈ 0.9013 (§7.2.6)")
rI_ref = r; r += 1
sc(ws3, r, 2, "Zona EUC")
sc(ws3, r, 3, None, fill=RES)
ws3.cell(r, 3).value = f'=IF(C{rI_ref}<0.3,"Insuficiente",IF(C{rI_ref}<0.5,"Básico",IF(C{rI_ref}<0.7,"Aceptable",IF(C{rI_ref}<0.85,"Bueno","Excelente"))))'
ws3.cell(r, 3).font = RF; r += 1
sc(ws3, r, 2, "Var(h) ≥ ε_min?")
sc(ws3, r, 3, None, fill=RES)
ws3.cell(r, 3).value = f'=IF(E{rvh}>=0.005,"✓ Var(h)≥0.005","✗ Piso no superado")'
ws3.cell(r, 3).font = RF; r += 3

# --- Coeficiente ρₕ (§7.2.4.1) ---
ttl(ws3, r, 1, "COEFICIENTE DE ADECUACIÓN ρₕ (§7.2.4.1)", 10); r += 1
note(ws3, r, 1, "Correlación Pearson entre h(x,Mᵤ) y satisfacción declarada s del usuario", 10); r += 1
hdr(ws3, r, ["", "Predicción", "h(x,Mᵤ)", "s (satisf.)", "", "", "", "", "", ""]); r += 1
rho_data = [
    (1, 0.27, 0.80), (2, 0.31, 0.85), (3, -0.05, 0.40), (4, 0.20, 0.70),
    (5, 0.41, 0.90), (6, -0.18, 0.25), (7, 0.32, 0.88), (8, 0.08, 0.55),
    (9, 0.34, 0.82), (10, 0.19, 0.65),
]
rho_start = r
for pred, hv, sv in rho_data:
    sc(ws3, r, 2, pred); sc(ws3, r, 3, hv, fill=INP, fmt='0.0000')
    sc(ws3, r, 4, sv, fill=INP, fmt='0.0000'); r += 1
rho_end = r - 1; r += 1

sc(ws3, r, 2, "ρₕ = CORREL(h, s)", font=SF)
sc(ws3, r, 3, None, fill=RES, fmt='0.0000')
ws3.cell(r, 3).value = f"=CORREL(C{rho_start}:C{rho_end},D{rho_start}:D{rho_end})"
ws3.cell(r, 3).font = RF
sc(ws3, r, 5, "Umbral ≥ 0.30")
sc(ws3, r, 7, None, fill=RES)
ws3.cell(r, 7).value = f'=IF(C{r}>=0.30,"✓ CUMPLE","✗ NO CUMPLE")'
ws3.cell(r, 7).font = RF
sc(ws3, r, 10, "V55 exige I≥0.70 AND ρₕ≥0.30 (§7.2.4.1)")


# ████████████████████████████████████████████████████████████████
# PESTAÑA 5: CAP 32 — TEORÍA LEY DEL AUMENTO
# ████████████████████████████████████████████████████████████████
ws4 = wb.create_sheet("Cap32_Teoria")
ws4.sheet_properties.tabColor = "E67E22"
setw(ws4, [4, 38, 35, 28, 20])
r = 1
ttl(ws4, r, 1, "CAPÍTULO 32 — MÉTRICAS PARA LA LEY DEL AUMENTO (§7.3)", 5); r += 2

ttl(ws4, r, 1, "1. FÓRMULAS PRINCIPALES (§7.3.3-7.3.8)", 5); r += 1
hdr(ws4, r, ["#", "Métrica", "Fórmula", "Significado", "Ref V55"]); r += 1
for i, (m, f, s, ref) in enumerate([
    ("Coef. Autonomía Preservada", "α(d) = α_base × (1 - M(d)) × Pᵤ", "Cuánta autonomía retiene usuario en decisión d", "§7.3.6"),
    ("Peso Moral M(d)", "AHP 7 indicadores → M(d) ∈ [0,1]", "Cuánta carga moral tiene la decisión", "§7.3.5"),
    ("Ratio Decisión Humana", "R = Σ[(1−α)×w] / Σ[w]  con w=1+M", "Agregado ponderado por moral", "§7.3.8"),
    ("Participación usuario", "Pᵤ ∈ [0,1]", "Nivel real de participación del usuario", "§7.3.7"),
], 1):
    sc(ws4, r, 1, i); sc(ws4, r, 2, m); sc(ws4, r, 3, f, font=FF); sc(ws4, r, 4, s); sc(ws4, r, 5, ref); r += 1
r += 1

# Tabla 52: 7 indicadores AHP para M(d)
ttl(ws4, r, 1, "2. INDICADORES PARA PESO MORAL M(d) — AHP (Tabla 52, §7.3.5)", 5); r += 1
hdr(ws4, r, ["#", "Indicador φ", "Descripción", "Escala", "Ref V55"]); r += 1
for i, (ind, desc, esc) in enumerate([
    ("φ₁: Reversibilidad", "¿La decisión es reversible?", "0=totalmente reversible, 1=irreversible"),
    ("φ₂: Alcance temporal", "¿Cuánto dura el efecto?", "0=efímero, 1=permanente"),
    ("φ₃: Personas afectadas", "¿Cuántas personas impacta?", "0=solo usuario, 1=muchas personas"),
    ("φ₄: Asimetría informativa", "¿El sistema sabe más que el usuario?", "0=equilibrio, 1=asimetría total"),
    ("φ₅: Vulnerabilidad", "¿El afectado es vulnerable?", "0=no vulnerable, 1=muy vulnerable"),
    ("φ₆: Dignidad afectada", "¿Afecta directamente a dignidad?", "0=no afecta, 1=afecta directamente"),
    ("φ₇: Marco normativo", "¿Existe regulación aplicable?", "0=no regulado, 1=altamente regulado"),
], 1):
    sc(ws4, r, 1, i); sc(ws4, r, 2, ind); sc(ws4, r, 3, desc); sc(ws4, r, 4, esc); sc(ws4, r, 5, "§7.3.5"); r += 1
r += 1

# Pesos AHP (Tablas 53-56)
ttl(ws4, r, 1, "3. PESOS AHP INDICADORES (Tablas 53-56, §7.3.5)", 5); r += 1
hdr(ws4, r, ["#", "Indicador", "Peso AHP wᵢ", "Justificación", ""]); r += 1
ahp_w = [("φ₁: Reversibilidad", 0.25), ("φ₂: Alcance temporal", 0.15),
          ("φ₃: Personas afectadas", 0.10), ("φ₄: Asimetría informativa", 0.15),
          ("φ₅: Vulnerabilidad", 0.15), ("φ₆: Dignidad afectada", 0.10),
          ("φ₇: Marco normativo", 0.10)]
for i, (ind, w) in enumerate(ahp_w, 1):
    sc(ws4, r, 1, i); sc(ws4, r, 2, ind); sc(ws4, r, 3, w, fill=WHT, fmt='0.00')
    sc(ws4, r, 4, "Derivado de matriz AHP (§7.3.5)"); r += 1
r_sum_w = r
sc(ws4, r, 2, "SUMA", font=SF); sc(ws4, r, 3, None, fill=CAL, fmt='0.00')
ws4.cell(r, 3).value = f"=SUM(C{r_sum_w-7}:C{r_sum_w-1})"; r += 2

# Tabla 60: HITL
ttl(ws4, r, 1, "4. MECANISMOS HITL POR ZONA DE RIESGO (Tabla 60, §7.3.14)", 5); r += 1
hdr(ws4, r, ["", "Zona", "M(d)", "Mecanismo HITL", ""]); r += 1
for zona, rango, mec, fl in [
    ("Baja", "M < 0.3", "Notificación post-hoc", GRN),
    ("Media", "0.3 ≤ M < 0.7", "Aprobación previa con explicación", YEL),
    ("Alta", "0.7 ≤ M < 0.9", "Doble aprobación + comité", ORG),
    ("Crítica", "M ≥ 0.9", "Abstención total del sistema", WRN),
]:
    sc(ws4, r, 2, zona, fill=fl); sc(ws4, r, 3, rango, fill=fl); sc(ws4, r, 4, mec, fill=fl); r += 1


# ████████████████████████████████████████████████████████████████
# PESTAÑA 6: CAP 32 — EJEMPLOS NUMÉRICOS
# ████████████████████████████████████████████████████████████████
ws5 = wb.create_sheet("Cap32_Ejemplos")
ws5.sheet_properties.tabColor = "D35400"
setw(ws5, [4, 28, 14, 14, 14, 14, 14, 14, 14, 30])
r = 1

ttl(ws5, r, 1, "EJEMPLOS NUMÉRICOS CAP 32 — Tablas 57-59 (§7.3.6-12)", 10); r += 2

scenarios = [
    ("Tabla 57: PRÉSTAMO HIPOTECARIO (§7.3.6)", [
        ("φ₁ Reversibilidad", 0.8, 0.25), ("φ₂ Alcance temporal", 0.9, 0.15),
        ("φ₃ Personas afectadas", 0.4, 0.10), ("φ₄ Asimetría info", 0.7, 0.15),
        ("φ₅ Vulnerabilidad", 0.6, 0.15), ("φ₆ Dignidad", 0.5, 0.10),
        ("φ₇ Marco normativo", 0.8, 0.10),
    ], 0.6, 0.7),
    ("Tabla 58: MODERACIÓN CONTENIDO MENORES (§7.3.9)", [
        ("φ₁ Reversibilidad", 0.3, 0.25), ("φ₂ Alcance temporal", 0.7, 0.15),
        ("φ₃ Personas afectadas", 0.8, 0.10), ("φ₄ Asimetría info", 0.6, 0.15),
        ("φ₅ Vulnerabilidad", 0.9, 0.15), ("φ₆ Dignidad", 0.7, 0.10),
        ("φ₇ Marco normativo", 0.7, 0.10),
    ], 0.6, 0.8),
    ("Tabla 59: DIAGNÓSTICO MÉDICO (§7.3.11)", [
        ("φ₁ Reversibilidad", 0.7, 0.25), ("φ₂ Alcance temporal", 0.8, 0.15),
        ("φ₃ Personas afectadas", 0.3, 0.10), ("φ₄ Asimetría info", 0.9, 0.15),
        ("φ₅ Vulnerabilidad", 0.8, 0.15), ("φ₆ Dignidad", 0.8, 0.10),
        ("φ₇ Marco normativo", 0.9, 0.10),
    ], 0.6, 0.5),
]

for title, indicators, alpha_base, pu in scenarios:
    ttl(ws5, r, 1, title, 10); r += 1
    hdr(ws5, r, ["#", "Indicador", "Valor φᵢ", "Peso wᵢ", "φᵢ×wᵢ", "", "", "", "", ""]); r += 1
    sc_start = r
    for i, (ind, phi, w) in enumerate(indicators, 1):
        sc(ws5, r, 1, i); sc(ws5, r, 2, ind)
        sc(ws5, r, 3, phi, fill=INP, fmt='0.00'); sc(ws5, r, 4, w, fill=WHT, fmt='0.00')
        sc(ws5, r, 5, None, fill=CAL, fmt='0.000')
        ws5.cell(r, 5).value = f"=C{r}*D{r}"; r += 1
    sc(ws5, r, 2, "M(d) = Σ(φᵢ×wᵢ)", font=SF)
    sc(ws5, r, 5, None, fill=RES, fmt='0.000')
    ws5.cell(r, 5).value = f"=SUM(E{sc_start}:E{r-1})"
    ws5.cell(r, 5).font = RF; rM = r; r += 1
    sc(ws5, r, 2, "α_base"); sc(ws5, r, 3, alpha_base, fill=INP, fmt='0.00'); r_ab = r; r += 1
    sc(ws5, r, 2, "Pᵤ (participación)"); sc(ws5, r, 3, pu, fill=INP, fmt='0.00'); r_pu = r; r += 1
    sc(ws5, r, 2, "α(d) = α_base×(1-M)×Pᵤ", font=SF)
    sc(ws5, r, 3, None, fill=RES, fmt='0.000')
    ws5.cell(r, 3).value = f"=C{r_ab}*(1-E{rM})*C{r_pu}"
    ws5.cell(r, 3).font = RF; r += 2

# Ratio R agregado (§7.3.8)
r += 1
ttl(ws5, r, 1, "RATIO R AGREGADO — 3 decisiones (§7.3.8)", 10); r += 1
hdr(ws5, r, ["#", "Decisión", "M(d)", "α_base", "Pᵤ", "α(d)", "(1−α)", "w=1+M", "(1−α)×w", ""]); r += 1
rd_start = r
dec_data = [
    ("Clasificar radiografía", 0.3, 0.6, 0.9),
    ("Sugerir diagnóstico", 0.7, 0.6, 0.7),
    ("Prescribir medicamento", 0.95, 0.6, 0.5),
]
for i, (dec, m, ab, pu) in enumerate(dec_data, 1):
    sc(ws5, r, 1, i); sc(ws5, r, 2, dec)
    sc(ws5, r, 3, m, fill=INP, fmt='0.00'); sc(ws5, r, 4, ab, fill=INP, fmt='0.00')
    sc(ws5, r, 5, pu, fill=INP, fmt='0.00')
    sc(ws5, r, 6, None, fill=CAL, fmt='0.000'); ws5.cell(r, 6).value = f"=D{r}*(1-C{r})*E{r}"
    sc(ws5, r, 7, None, fill=CAL, fmt='0.000'); ws5.cell(r, 7).value = f"=1-F{r}"
    sc(ws5, r, 8, None, fill=CAL, fmt='0.00'); ws5.cell(r, 8).value = f"=1+C{r}"
    sc(ws5, r, 9, None, fill=CAL, fmt='0.000'); ws5.cell(r, 9).value = f"=G{r}*H{r}"
    r += 1
rd_end = r - 1
sc(ws5, r, 2, "SUMA", font=SF)
sc(ws5, r, 8, None, fill=CAL, fmt='0.000'); ws5.cell(r, 8).value = f"=SUM(H{rd_start}:H{rd_end})"
sc(ws5, r, 9, None, fill=CAL, fmt='0.000'); ws5.cell(r, 9).value = f"=SUM(I{rd_start}:I{rd_end})"
rsum = r; r += 2
sc(ws5, r, 2, "R = Σ[(1−α)×w] / Σ[w]", font=SF)
sc(ws5, r, 3, None, fill=RES, fmt='0.0000'); ws5.cell(r, 3).value = f"=I{rsum}/H{rsum}"
ws5.cell(r, 3).font = RF
sc(ws5, r, 5, "Umbral ≥ 0.70")
sc(ws5, r, 6, None, fill=RES)
ws5.cell(r, 6).value = f'=IF(C{r}>=0.70,"✓ CUMPLE","✗ NO CUMPLE")'
ws5.cell(r, 6).font = RF
rR_ref = r


# ████████████████████████████████████████████████████████████████
# PESTAÑA 7: CAP 33 — TEORÍA LEY DE LA TRANSPARENCIA
# ████████████████████████████████████████████████████████████████
ws6 = wb.create_sheet("Cap33_Teoria")
ws6.sheet_properties.tabColor = "E74C3C"
setw(ws6, [4, 38, 38, 28, 20])
r = 1
ttl(ws6, r, 1, "CAPÍTULO 33 — MÉTRICAS PARA LA LEY DE LA TRANSPARENCIA (§7.4)", 5); r += 2

ttl(ws6, r, 1, "1. MÉTRICAS PRINCIPALES (§7.4.1-7.4.10)", 5); r += 1
hdr(ws6, r, ["#", "Métrica", "Fórmula", "Umbral", "Ref V55"]); r += 1
for i, (m, f, u, ref) in enumerate([
    ("Cobertura Explicación", "CE = decisiones_con_SHAP / total", "CE ≥ 0.90", "§7.4.3"),
    ("Índ. Calidad Explicación", "ICE = (Q_F + Q_C + Q_Ac) / 3", "[0,1]", "§7.4.5"),
    ("Cobertura Efectiva Expl.", "CEE = CE × ICE", "CEE ≥ 0.50 (piso)", "§7.4.5"),
    ("Índ. Control Datos", "ICD = (Cₐ+Cc+Cp+Cr+Cₑ)/5", "ICD ≥ 0.40 (piso)", "§7.4.6"),
    ("Índ. Distribución Poder", "IDP = (Dₐ+Ds+Dᵢ+Dᵥ)/4", "IDP ≥ 0.25 (piso)", "§7.4.8"),
    ("Índ. Compuesto Transp.", "ICT = wₑ×CEE + wᵢ×ICD + wₚ×IDP", "ICT ≥ 0.75", "§7.4.11"),
], 1):
    sc(ws6, r, 1, i); sc(ws6, r, 2, m); sc(ws6, r, 3, f, font=FF); sc(ws6, r, 4, u); sc(ws6, r, 5, ref); r += 1
r += 1

# Pesos ICT (Tablas 69-73)
ttl(ws6, r, 1, "2. PESOS AHP PARA ICT (Tablas 69-73, §7.4.11-12)", 5); r += 1
hdr(ws6, r, ["#", "Componente", "Peso wᵢ", "Justificación", ""]); r += 1
for i, (comp, w, just) in enumerate([
    ("CEE (Explicabilidad)", 0.50, "Mayor peso: transparencia algorítmica fundamental"),
    ("ICD (Control de datos)", 0.30, "RGPD y derechos de datos"),
    ("IDP (Distribución poder)", 0.20, "Concentración informacional"),
], 1):
    sc(ws6, r, 1, i); sc(ws6, r, 2, comp); sc(ws6, r, 3, w, fill=WHT, fmt='0.00'); sc(ws6, r, 4, just); r += 1
r += 1

# Tabla 64: SHAP ejemplo préstamo
ttl(ws6, r, 1, "3. EJEMPLO SHAP — Préstamo (Tabla 64, §7.4.4)", 5); r += 1
hdr(ws6, r, ["#", "Variable", "Valor SHAP", "Interpretación", ""]); r += 1
for i, (v, shap, interp) in enumerate([
    ("Ingresos anuales", +0.15, "Aumenta probabilidad de aprobación"),
    ("Historial crediticio", +0.22, "Factor más positivo"),
    ("Ratio deuda/ingreso", -0.18, "Reduce probabilidad"),
    ("Antigüedad laboral", +0.08, "Contribución positiva moderada"),
    ("Edad", -0.03, "Impacto mínimo"),
], 1):
    sc(ws6, r, 1, i); sc(ws6, r, 2, v); sc(ws6, r, 3, shap, fill=LIL, fmt='+0.00;-0.00')
    sc(ws6, r, 4, interp); r += 1


# ████████████████████████████████████████████████████████████████
# PESTAÑA 8: CAP 33 — EJEMPLOS NUMÉRICOS
# ████████████████████████████████████████████████████████████████
ws7 = wb.create_sheet("Cap33_Ejemplos")
ws7.sheet_properties.tabColor = "C0392B"
setw(ws7, [4, 30, 14, 14, 14, 14, 14, 30])
r = 1
ttl(ws7, r, 1, "EJEMPLOS NUMÉRICOS CAP 33 — ICT COMPLETO (§7.4.11-15)", 8); r += 2

# CEE
ttl(ws7, r, 1, "1. COBERTURA EFECTIVA DE EXPLICACIÓN — CEE", 8); r += 1
sc(ws7, r, 2, "CE (Cobertura)"); sc(ws7, r, 3, 0.92, fill=INP, fmt='0.00')
sc(ws7, r, 8, "92% decisiones con SHAP"); rCE = r; r += 1
sc(ws7, r, 2, "Q_F (Fidelidad)"); sc(ws7, r, 3, 0.85, fill=INP, fmt='0.00'); rQF = r; r += 1
sc(ws7, r, 2, "Q_C (Comprensión)"); sc(ws7, r, 3, 0.70, fill=INP, fmt='0.00'); rQC = r; r += 1
sc(ws7, r, 2, "Q_Ac (Accionabilidad)"); sc(ws7, r, 3, 0.65, fill=INP, fmt='0.00'); rQA = r; r += 1
sc(ws7, r, 2, "ICE = (Q_F+Q_C+Q_Ac)/3", font=SF)
sc(ws7, r, 3, None, fill=CAL, fmt='0.0000')
ws7.cell(r, 3).value = f"=(C{rQF}+C{rQC}+C{rQA})/3"; rICE = r; r += 1
sc(ws7, r, 2, "CEE = CE × ICE", font=SF)
sc(ws7, r, 3, None, fill=RES, fmt='0.0000')
ws7.cell(r, 3).value = f"=C{rCE}*C{rICE}"; ws7.cell(r, 3).font = RF
sc(ws7, r, 5, "Piso ≥ 0.50")
sc(ws7, r, 6, None, fill=RES)
ws7.cell(r, 6).value = f'=IF(C{r}>=0.50,"✓","✗")'
rCEE = r; r += 2

# ICD
ttl(ws7, r, 1, "2. ÍNDICE DE CONTROL DE DATOS — ICD (§7.4.6)", 8); r += 1
hdr(ws7, r, ["", "Derecho RGPD", "Puntuación Cᵢ", "", "", "", "", ""]); r += 1
icd_items = [("Cₐ: Acceso", 0.80), ("Cc: Corrección", 0.75), ("Cp: Portabilidad", 0.60),
             ("Cr: Restricción", 0.70), ("Cₑ: Eliminación", 0.65)]
icd_start = r
for item, val in icd_items:
    sc(ws7, r, 2, item); sc(ws7, r, 3, val, fill=INP, fmt='0.00'); r += 1
icd_end = r - 1
sc(ws7, r, 2, "ICD = media", font=SF)
sc(ws7, r, 3, None, fill=RES, fmt='0.0000')
ws7.cell(r, 3).value = f"=AVERAGE(C{icd_start}:C{icd_end})"
ws7.cell(r, 3).font = RF
sc(ws7, r, 5, "Piso ≥ 0.40")
sc(ws7, r, 6, None, fill=RES)
ws7.cell(r, 6).value = f'=IF(C{r}>=0.40,"✓","✗")'
rICD = r; r += 2

# IDP
ttl(ws7, r, 1, "3. ÍNDICE DE DISTRIBUCIÓN DE PODER — IDP (§7.4.8)", 8); r += 1
hdr(ws7, r, ["", "Dimensión", "Puntuación Dᵢ", "", "", "", "", ""]); r += 1
idp_items = [("Dₐ: Acceso algoritmos", 0.55), ("Ds: Simetría info", 0.50),
             ("Dᵢ: Interoperabilidad", 0.45), ("Dᵥ: Voz efectiva", 0.60)]
idp_start = r
for item, val in idp_items:
    sc(ws7, r, 2, item); sc(ws7, r, 3, val, fill=INP, fmt='0.00'); r += 1
idp_end = r - 1
sc(ws7, r, 2, "IDP = media", font=SF)
sc(ws7, r, 3, None, fill=RES, fmt='0.0000')
ws7.cell(r, 3).value = f"=AVERAGE(C{idp_start}:C{idp_end})"
ws7.cell(r, 3).font = RF
sc(ws7, r, 5, "Piso ≥ 0.25")
sc(ws7, r, 6, None, fill=RES)
ws7.cell(r, 6).value = f'=IF(C{r}>=0.25,"✓","✗")'
rIDP = r; r += 2

# ICT = ponderado
ttl(ws7, r, 1, "4. ÍNDICE COMPUESTO DE TRANSPARENCIA — ICT (Tabla 74, §7.4.15)", 8); r += 1
hdr(ws7, r, ["", "Componente", "Valor", "Peso", "Ponderado", "", "", ""]); r += 1
sc(ws7, r, 2, "CEE"); sc(ws7, r, 3, None, fill=CAL, fmt='0.0000')
ws7.cell(r, 3).value = f"=C{rCEE}"; sc(ws7, r, 4, 0.50, fill=WHT, fmt='0.00')
sc(ws7, r, 5, None, fill=CAL, fmt='0.0000'); ws7.cell(r, 5).value = f"=C{r}*D{r}"; r_ict1 = r; r += 1
sc(ws7, r, 2, "ICD"); sc(ws7, r, 3, None, fill=CAL, fmt='0.0000')
ws7.cell(r, 3).value = f"=C{rICD}"; sc(ws7, r, 4, 0.30, fill=WHT, fmt='0.00')
sc(ws7, r, 5, None, fill=CAL, fmt='0.0000'); ws7.cell(r, 5).value = f"=C{r}*D{r}"; r_ict2 = r; r += 1
sc(ws7, r, 2, "IDP"); sc(ws7, r, 3, None, fill=CAL, fmt='0.0000')
ws7.cell(r, 3).value = f"=C{rIDP}"; sc(ws7, r, 4, 0.20, fill=WHT, fmt='0.00')
sc(ws7, r, 5, None, fill=CAL, fmt='0.0000'); ws7.cell(r, 5).value = f"=C{r}*D{r}"; r_ict3 = r; r += 1
sc(ws7, r, 2, "ICT = Σ ponderados", font=SF)
sc(ws7, r, 5, None, fill=RES, fmt='0.0000')
ws7.cell(r, 5).value = f"=E{r_ict1}+E{r_ict2}+E{r_ict3}"
ws7.cell(r, 5).font = RF
sc(ws7, r, 7, None, fill=RES)
ws7.cell(r, 7).value = f'=IF(E{r}>=0.75,"✓ CUMPLE","✗ NO CUMPLE")'
ws7.cell(r, 7).font = RF
rICT_ref = r


# ████████████████████████████████████████████████████████████████
# PESTAÑA 9: CAP 34 — TEORÍA LEY DEL LÍMITE
# ████████████████████████████████████████████████████████████████
ws8 = wb.create_sheet("Cap34_Teoria")
ws8.sheet_properties.tabColor = "8E44AD"
setw(ws8, [4, 38, 38, 28, 20])
r = 1
ttl(ws8, r, 1, "CAPÍTULO 34 — MÉTRICAS PARA LA LEY DEL LÍMITE (§7.5)", 5); r += 2

ttl(ws8, r, 1, "1. CLASIFICADOR DE DECISIONES MORALES (Tabla 76, §7.5.8)", 5); r += 1
hdr(ws8, r, ["#", "Componente", "Fórmula", "Significado", "Ref V55"]); r += 1
for i, (c, f, s, ref) in enumerate([
    ("Probabilidad moral", "P(moral|d) = σ(Σ wᵢφᵢ − θ)", "Sigmoide sobre indicadores ponderados", "§7.5.8"),
    ("Función sigmoide", "σ(z) = 1 / (1 + e⁻ᶻ)", "Mapea score a probabilidad [0,1]", "§7.5.8"),
    ("Umbral decisión", "θ = umbral calibrado", "θ=3.5 por defecto", "§7.5.9"),
    ("Tasa Abstención", "A = VP / (VP + FN)", "Recall: morales detectadas / total morales", "§7.5.17"),
], 1):
    sc(ws8, r, 1, i); sc(ws8, r, 2, c); sc(ws8, r, 3, f, font=FF); sc(ws8, r, 4, s); sc(ws8, r, 5, ref); r += 1
r += 1

# Tabla 77+81: Indicadores φ con pesos
ttl(ws8, r, 1, "2. INDICADORES φ Y PESOS (Tablas 77, 81, §7.5.7-13)", 5); r += 1
hdr(ws8, r, ["#", "Indicador", "Peso wᵢ", "Escala", "Ref V55"]); r += 1
phi_w = [("φ₁: Reversibilidad", 1.0, "0-1"), ("φ₂: Pluralismo razonable", 1.5, "0-1"),
         ("φ₃: Afectación derechos", 1.2, "0-1"), ("φ₄: Vulnerabilidad", 1.3, "0-1"),
         ("φ₅: Incertidumbre empírica", 0.8, "0-1"), ("φ₆: Vida o muerte", 2.0, "0-1")]
for i, (ind, w, esc) in enumerate(phi_w, 1):
    sc(ws8, r, 1, i); sc(ws8, r, 2, ind); sc(ws8, r, 3, w, fill=WHT, fmt='0.0')
    sc(ws8, r, 4, esc); sc(ws8, r, 5, "§7.5.13"); r += 1


# ████████████████████████████████████████████████████████████████
# PESTAÑA 10: CAP 34 — EJEMPLOS NUMÉRICOS
# ████████████████████████████████████████████████████████████████
ws9 = wb.create_sheet("Cap34_Ejemplos")
ws9.sheet_properties.tabColor = "7D3C98"
setw(ws9, [4, 28, 10, 10, 10, 10, 10, 10, 14, 14, 14, 20])
r = 1
ttl(ws9, r, 1, "EJEMPLOS NUMÉRICOS CAP 34 (§7.5.7-26)", 12); r += 2

# Tabla 78-80: 3 decisiones
ttl(ws9, r, 1, "1. TRES DECISIONES — P(moral|d) (Tablas 78-80, §7.5.14-16)", 12); r += 1
hdr(ws9, r, ["", "Decisión", "φ₁", "φ₂", "φ₃", "φ₄", "φ₅", "φ₆", "Score z", "P(moral)", "Umbral θ", "Dictamen"]); r += 1
decisions_34 = [
    ("Recomendar producto", 0.1, 0.1, 0.2, 0.1, 0.2, 0.0),
    ("Denegar préstamo", 0.7, 0.5, 0.8, 0.6, 0.4, 0.0),
    ("Triaje emergencia", 0.9, 0.8, 0.9, 0.9, 0.7, 1.0),
]
theta = 3.5
weights_34 = [1.0, 1.5, 1.2, 1.3, 0.8, 2.0]
for dec, *phis in decisions_34:
    sc(ws9, r, 2, dec)
    for j, phi in enumerate(phis):
        sc(ws9, r, 3+j, phi, fill=INP, fmt='0.0')
    z_val = sum(w*p for w, p in zip(weights_34, phis)) - theta
    p_val = 1 / (1 + math.exp(-z_val))
    sc(ws9, r, 9, round(z_val, 3), fill=CAL, fmt='0.000')
    sc(ws9, r, 10, round(p_val, 3), fill=CAL, fmt='0.000')
    sc(ws9, r, 11, theta, fill=WHT)
    sc(ws9, r, 12, None, fill=RES)
    if p_val >= 0.95: ws9.cell(r, 12).value = "⛔ ABSTENCIÓN"
    elif p_val >= 0.5: ws9.cell(r, 12).value = "⚠ Supervisión"
    else: ws9.cell(r, 12).value = "✓ Técnica"
    ws9.cell(r, 12).font = RF; r += 1
r += 2

# Tabla 82-83: Espectro 9 decisiones
ttl(ws9, r, 1, "2. ESPECTRO DE 9 DECISIONES (Tablas 82-83, §7.5.18)", 12); r += 1
hdr(ws9, r, ["#", "Decisión", "φ₁", "φ₂", "φ₃", "φ₄", "φ₅", "φ₆", "Score z", "P(moral)", "", "Dictamen"]); r += 1
spectrum = [
    ("Ordenar playlist", 0.05, 0.05, 0.05, 0.05, 0.1, 0.0),
    ("Filtrar spam", 0.1, 0.1, 0.1, 0.1, 0.1, 0.0),
    ("Recomendar noticia", 0.2, 0.3, 0.2, 0.2, 0.2, 0.0),
    ("Priorizar CV", 0.5, 0.4, 0.6, 0.3, 0.3, 0.0),
    ("Denegar seguro", 0.7, 0.5, 0.7, 0.6, 0.4, 0.0),
    ("Moderar contenido", 0.4, 0.6, 0.5, 0.7, 0.5, 0.0),
    ("Scoring crediticio", 0.7, 0.5, 0.8, 0.6, 0.4, 0.0),
    ("Diagnóstico médico", 0.8, 0.6, 0.8, 0.8, 0.6, 0.3),
    ("Arma autónoma", 1.0, 0.9, 1.0, 1.0, 0.9, 1.0),
]
for idx, (dec, *phis) in enumerate(spectrum, 1):
    sc(ws9, r, 1, idx); sc(ws9, r, 2, dec)
    for j, phi in enumerate(phis):
        sc(ws9, r, 3+j, phi, fill=INP, fmt='0.0')
    z_val = sum(w*p for w, p in zip(weights_34, phis)) - theta
    p_val = 1 / (1 + math.exp(-z_val))
    sc(ws9, r, 9, round(z_val, 3), fill=CAL, fmt='0.000')
    sc(ws9, r, 10, round(p_val, 3), fill=CAL, fmt='0.000')
    if p_val >= 0.95: dic = "⛔ ABSTENCIÓN"
    elif p_val >= 0.5: dic = "⚠ Supervisión"
    else: dic = "✓ Técnica"
    sc(ws9, r, 12, dic, fill=RES); ws9.cell(r, 12).font = RF; r += 1
r += 2

# Tasa de Abstención A (§7.5.17)
ttl(ws9, r, 1, "3. TASA DE ABSTENCIÓN CORRECTA A (§7.5.17-19)", 12); r += 1
sc(ws9, r, 2, "Total decisiones auditadas"); sc(ws9, r, 3, 100, fill=INP); rn = r; r += 1
sc(ws9, r, 2, "Genuinamente morales (panel)"); sc(ws9, r, 3, 20, fill=INP); rgm = r; r += 1
sc(ws9, r, 2, "Genuinamente técnicas"); sc(ws9, r, 3, None, fill=CAL)
ws9.cell(r, 3).value = f"=C{rn}-C{rgm}"; r += 1
sc(ws9, r, 2, "VP (morales detectadas)"); sc(ws9, r, 3, 19, fill=INP); rvp = r; r += 1
sc(ws9, r, 2, "FN (morales NO detectadas)"); sc(ws9, r, 3, None, fill=CAL)
ws9.cell(r, 3).value = f"=C{rgm}-C{rvp}"; rfn = r; r += 1
sc(ws9, r, 2, "FP (abstención innecesaria)"); sc(ws9, r, 3, 15, fill=INP); r += 2
sc(ws9, r, 2, "A = VP / (VP + FN)", font=SF)
sc(ws9, r, 3, None, fill=RES, fmt='0.0000')
ws9.cell(r, 3).value = f"=C{rvp}/(C{rvp}+C{rfn})"
ws9.cell(r, 3).font = RF
sc(ws9, r, 5, "Umbral ≥ 0.95")
sc(ws9, r, 6, None, fill=RES)
ws9.cell(r, 6).value = f'=IF(C{r}>=0.95,"✓ CUMPLE","✗ NO CUMPLE")'
ws9.cell(r, 6).font = RF
rA_ref = r


# ████████████████████████████████████████████████████████████████
# PESTAÑA 11: CAP 35 — IHG + CONDICIONES
# ████████████████████████████████████████████████████████████████
ws10 = wb.create_sheet("Cap35_IHG")
ws10.sheet_properties.tabColor = "D4AC0D"
setw(ws10, [4, 32, 14, 14, 18, 18, 30])
r = 1
ttl(ws10, r, 1, "CAPÍTULO 35 — ÍNDICE DE HUMANIZACIÓN GLOBAL (§7.6.4)", 7); r += 1
note(ws10, r, 1, "IHG = I^w₁ × R^w₂ × ICT^w₃ × A^w₄  con w=[0.35, 0.30, 0.20, 0.15]", 7); r += 2

# Pesos AHP (Tablas 89-91)
ttl(ws10, r, 1, "PESOS AHP PARA LAS LEYES (Tablas 89-91, §7.6.3-4)", 7); r += 1
hdr(ws10, r, ["", "Métrica (Ley)", "Valor", "Peso wᵢ", "Métricaʷ", "Contribución", "Nota"]); r += 1
metrics_ihg = [
    ("I — Persona (Cap.31)", 0.80, 0.35),
    ("R — Aumento (Cap.32)", 0.75, 0.30),
    ("ICT — Transparencia (Cap.33)", 0.72, 0.20),
    ("A — Límite (Cap.34)", 0.95, 0.15),
]
dm = []
for m, v, w in metrics_ihg:
    sc(ws10, r, 2, m); sc(ws10, r, 3, v, fill=INP, fmt='0.0000')
    sc(ws10, r, 4, w, fill=WHT, fmt='0.00')
    sc(ws10, r, 5, None, fill=CAL, fmt='0.0000')
    ws10.cell(r, 5).value = f"=C{r}^D{r}"
    sc(ws10, r, 6, None, fill=CAL, fmt='0.0000')
    ws10.cell(r, 6).value = f"=D{r}*LN(C{r})"
    dm.append(r); r += 1
r += 1

# IHG
sc(ws10, r, 2, "IHG = Π(mᵢ^wᵢ)", font=SF)
sc(ws10, r, 3, None, fill=RES, fmt='0.0000')
prod = "*".join([f"E{d}" for d in dm])
ws10.cell(r, 3).value = f"={prod}"
ws10.cell(r, 3).font = RF
rIHG_d = r; r += 1
sc(ws10, r, 2, "Zona EUC")
sc(ws10, r, 3, None, fill=RES)
ws10.cell(r, 3).value = f'=IF(C{rIHG_d}<0.5,"Crítico/Insuficiente",IF(C{rIHG_d}<0.6,"Aceptable bajo",IF(C{rIHG_d}<0.7,"Aceptable",IF(C{rIHG_d}<0.85,"Bueno","Excelente"))))'
ws10.cell(r, 3).font = RF; r += 2

# Tabla 92: Umbrales IHG
ttl(ws10, r, 1, "UMBRALES DE IHG PROPUESTOS (Tabla 92, §7.6.7)", 7); r += 1
hdr(ws10, r, ["", "Nivel", "Rango IHG", "Significado", "", "", ""]); r += 1
for niv, rng, sig, fl in [
    ("No certificable", "IHG < 0.60", "No supera umbral mínimo", WRN),
    ("Certificación básica", "0.60 ≤ IHG < 0.70", "Cumplimiento mínimo aceptable", YEL),
    ("Certificación estándar", "0.70 ≤ IHG < 0.85", "Cumplimiento sólido", GRN),
    ("Certificación excelente", "IHG ≥ 0.85", "Referencia del sector", GRN2),
]:
    sc(ws10, r, 2, niv, fill=fl); sc(ws10, r, 3, rng, fill=fl); sc(ws10, r, 4, sig, fill=fl); r += 1
r += 1

# 3 Condiciones necesarias (§7.6.8)
ttl(ws10, r, 1, "3 CONDICIONES NECESARIAS (§7.6.8)", 7); r += 1
sc(ws10, r, 2, "C1: Pisos mínimos (I≥0.50, R≥0.50, ICT≥0.50, A≥0.80)")
sc(ws10, r, 3, None, fill=RES)
ws10.cell(r, 3).value = f'=IF(AND(C{dm[0]}>=0.5,C{dm[1]}>=0.5,C{dm[2]}>=0.5,C{dm[3]}>=0.8),"✓","✗")'
ws10.cell(r, 3).font = RF; r += 1
sc(ws10, r, 2, "C2: IHG ≥ 0.60")
sc(ws10, r, 3, None, fill=RES)
ws10.cell(r, 3).value = f'=IF(C{rIHG_d}>=0.6,"✓","✗")'
ws10.cell(r, 3).font = RF; r += 1
sc(ws10, r, 2, "C3: Desequilibrio ≤ 0.40")
sc(ws10, r, 3, None, fill=RES)
ws10.cell(r, 3).value = f'=IF((MAX(C{dm[0]},C{dm[1]},C{dm[2]},C{dm[3]})-MIN(C{dm[0]},C{dm[1]},C{dm[2]},C{dm[3]}))<=0.4,"✓","✗")'
ws10.cell(r, 3).font = RF


# ████████████████████████████████████████████████████████████████
# PESTAÑA 12: CAP 35 — VALIDACIÓN Y SENSIBILIDAD
# ████████████████████████████████████████████████████████████████
ws11 = wb.create_sheet("Cap35_Validacion")
ws11.sheet_properties.tabColor = "F39C12"
setw(ws11, [4, 30, 14, 14, 14, 14, 14, 30])
r = 1
ttl(ws11, r, 1, "CAP 35 — SENSIBILIDAD, PATOLOGÍAS, EROSIÓN (§7.6.10-14)", 8); r += 2

# Tabla 93: Patologías
ttl(ws11, r, 1, "1. PATOLOGÍAS DE DISEÑO (Tabla 93, §7.6.10)", 8); r += 1
hdr(ws11, r, ["#", "Patología", "Perfil", "Diagnóstico", "", "", "", ""]); r += 1
for i, (pat, perf, diag) in enumerate([
    ("Paternalismo algorítmico", "I alto, R bajo", "Conoce al usuario pero decide por él"),
    ("Transparencia vacía", "ICT alto, I bajo", "Explica pero no personaliza"),
    ("Caja negra benevolente", "I+R altos, ICT bajo", "Funciona bien pero no es auditable"),
    ("Límite sin persona", "A alto, I bajo", "Se abstiene correctamente sin conocer al usuario"),
    ("Cumplimiento cosmético", "Todas ~0.55", "Mínimos sin excelencia en ninguna dimensión"),
], 1):
    sc(ws11, r, 1, i); sc(ws11, r, 2, pat); sc(ws11, r, 3, perf, fill=LIL); sc(ws11, r, 4, diag); r += 1
r += 1

# Tabla 94: Sensibilidad pesos
ttl(ws11, r, 1, "2. SENSIBILIDAD DE PESOS ±10% (Tabla 94, §7.6.13)", 8); r += 1
hdr(ws11, r, ["#", "Escenario", "w_I", "w_R", "w_ICT", "w_A", "IHG", "Nota"]); r += 1
sens_scenarios = [
    ("Base", 0.35, 0.30, 0.20, 0.15),
    ("w_I+10%", 0.385, 0.2717, 0.1811, 0.1622),
    ("w_R+10%", 0.3182, 0.33, 0.1855, 0.1664),
    ("w_ICT+10%", 0.3208, 0.2750, 0.22, 0.1842),
    ("w_A+10%", 0.3231, 0.2769, 0.1846, 0.1654),
]
for i, (esc, wi, wr, wt, wa) in enumerate(sens_scenarios, 1):
    sc(ws11, r, 1, i); sc(ws11, r, 2, esc)
    sc(ws11, r, 3, wi, fill=INP if i > 1 else WHT, fmt='0.000')
    sc(ws11, r, 4, wr, fill=INP if i > 1 else WHT, fmt='0.000')
    sc(ws11, r, 5, wt, fill=INP if i > 1 else WHT, fmt='0.000')
    sc(ws11, r, 6, wa, fill=INP if i > 1 else WHT, fmt='0.000')
    ihg_val = (0.80**wi) * (0.75**wr) * (0.72**wt) * (0.95**wa)
    sc(ws11, r, 7, round(ihg_val, 4), fill=RES, fmt='0.0000')
    ws11.cell(r, 7).font = RF; r += 1
r += 1

# Erosión temporal (§7.6.12)
ttl(ws11, r, 1, "3. EROSIÓN TEMPORAL (§7.6.12)", 8); r += 1
note(ws11, r, 1, "Monitoreo: si IHG(t) cae >5% respecto a certificación → alerta", 8); r += 1
sc(ws11, r, 2, "IHG certificación"); sc(ws11, r, 3, 0.79, fill=INP, fmt='0.0000'); rC = r; r += 1
sc(ws11, r, 2, "IHG actual"); sc(ws11, r, 3, 0.74, fill=INP, fmt='0.0000'); rA = r; r += 1
sc(ws11, r, 2, "Caída %"); sc(ws11, r, 3, None, fill=CAL, fmt='0.00%')
ws11.cell(r, 3).value = f"=(C{rC}-C{rA})/C{rC}"; r += 1
sc(ws11, r, 2, "¿Alerta erosión?"); sc(ws11, r, 3, None, fill=RES)
ws11.cell(r, 3).value = f'=IF((C{rC}-C{rA})/C{rC}>0.05,"⚠ ALERTA EROSIÓN","✓ Estable")'
ws11.cell(r, 3).font = RF


# ████████████████████████████████████████████████████████████████
# PESTAÑA 13: DASHBOARD — PANEL RESUMEN
# ████████████████████████████████████████████████████████████████
ws12 = wb.create_sheet("Dashboard")
ws12.sheet_properties.tabColor = "1ABC9C"
setw(ws12, [4, 32, 14, 14, 18, 30])
r = 1
ttl(ws12, r, 1, "═══ DASHBOARD — PANEL DE CERTIFICACIÓN V55 ═══", 6); r += 1
note(ws12, r, 1, "Dr. Óscar Gallego Castilla — Las Cuatro Leyes de la IA Humana", 6); r += 2

# Nivel 1: D1-D4
ttl(ws12, r, 1, "NIVEL 1: CONDICIONES DEONTOLÓGICAS", 6); r += 1
hdr(ws12, r, ["", "Condición", "Cumple (1/0)", "Resultado", "", ""]); r += 1
dd_rows = []
for d in ["D1: No instrumentalización", "D2: Override humano",
          "D3: Explicación accesible", "D4: Abstención moral"]:
    sc(ws12, r, 2, d); sc(ws12, r, 3, 1, fill=INP)
    sc(ws12, r, 4, None, fill=RES)
    ws12.cell(r, 4).value = f'=IF(C{r}=1,"✓","✗ DESCALIFICADO")'
    ws12.cell(r, 4).font = RF; dd_rows.append(r); r += 1
sc(ws12, r, 2, "NIVEL 1", font=SF); sc(ws12, r, 3, None, fill=RES)
dd_and = "*".join([f"C{d}" for d in dd_rows])
ws12.cell(r, 3).value = f'=IF({dd_and}=1,"✓ SUPERADO","✗ DESCALIFICADO")'
ws12.cell(r, 3).font = RF; r += 2

# Nivel 2: Métricas
ttl(ws12, r, 1, "NIVEL 2: MÉTRICAS POR LEY", 6); r += 1
hdr(ws12, r, ["", "Métrica", "Valor", "Umbral", "EUC", ""]); r += 1
dm2 = []
for met, val, umb in [
    ("I — Persona", 0.80, "≥ 0.70"),
    ("R — Aumento", 0.75, "≥ 0.70"),
    ("ICT — Transparencia", 0.72, "≥ 0.75"),
    ("A — Límite", 0.95, "≥ 0.95"),
]:
    sc(ws12, r, 2, met); sc(ws12, r, 3, val, fill=INP, fmt='0.0000')
    sc(ws12, r, 4, umb)
    sc(ws12, r, 5, None, fill=RES)
    ws12.cell(r, 5).value = f'=IF(C{r}<0.3,"Crítico",IF(C{r}<0.5,"Insuficiente",IF(C{r}<0.7,"Aceptable",IF(C{r}<0.85,"Bueno","Excelente"))))'
    ws12.cell(r, 5).font = RF; dm2.append(r); r += 1
r += 1

# IHG
ttl(ws12, r, 1, "IHG — ÍNDICE DE HUMANIZACIÓN GLOBAL", 6); r += 1
sc(ws12, r, 2, "IHG = I^0.35 × R^0.30 × ICT^0.20 × A^0.15", font=SF)
sc(ws12, r, 3, None, fill=RES, fmt='0.0000')
ws12.cell(r, 3).value = f"=C{dm2[0]}^0.35*C{dm2[1]}^0.30*C{dm2[2]}^0.20*C{dm2[3]}^0.15"
ws12.cell(r, 3).font = Font(name='Calibri', size=14, bold=True, color='006100')
rIHG_d = r; r += 2

# 3 Condiciones
ttl(ws12, r, 1, "3 CONDICIONES NECESARIAS (§7.6.8)", 6); r += 1
sc(ws12, r, 2, "C1: Pisos mínimos (I≥0.50, R≥0.50, ICT≥0.50, A≥0.80)")
sc(ws12, r, 3, None, fill=RES)
ws12.cell(r, 3).value = f'=IF(AND(C{dm2[0]}>=0.5,C{dm2[1]}>=0.5,C{dm2[2]}>=0.5,C{dm2[3]}>=0.8),"✓","✗")'
ws12.cell(r, 3).font = RF; r += 1
sc(ws12, r, 2, "C2: IHG ≥ 0.60")
sc(ws12, r, 3, None, fill=RES)
ws12.cell(r, 3).value = f'=IF(C{rIHG_d}>=0.6,"✓","✗")'
ws12.cell(r, 3).font = RF; r += 1
sc(ws12, r, 2, "C3: Desequilibrio ≤ 0.40")
sc(ws12, r, 3, None, fill=RES)
ws12.cell(r, 3).value = f'=IF((MAX(C{dm2[0]},C{dm2[1]},C{dm2[2]},C{dm2[3]})-MIN(C{dm2[0]},C{dm2[1]},C{dm2[2]},C{dm2[3]}))<=0.4,"✓","✗")'
ws12.cell(r, 3).font = RF; r += 2

# DICTAMEN FINAL
ttl(ws12, r, 1, "═══ DICTAMEN FINAL ═══", 6); r += 1
sc(ws12, r, 2, "RESULTADO CERTIFICACIÓN", font=Font(name='Calibri', size=14, bold=True))
sc(ws12, r, 3, None, fill=RES)
ws12.cell(r, 3).font = Font(name='Calibri', size=14, bold=True, color='006100')
ws12.cell(r, 3).value = f'=IF(C{rIHG_d}>=0.85,"EXCELENTE",IF(C{rIHG_d}>=0.70,"ESTÁNDAR",IF(C{rIHG_d}>=0.60,"BÁSICA","NO CERTIFICABLE")))'


# ════════════════════════════════════════════════════════════════
# GUARDAR
# ════════════════════════════════════════════════════════════════
fn = "V55_Modelo_Completo_v2.xlsx"
wb.save(fn)
print(f"""
╔══════════════════════════════════════════════════════════════╗
║  ✅ GENERADO: {fn}                         ║
╠══════════════════════════════════════════════════════════════╣
║  13 pestañas — Capítulos 30 a 35 + Dashboard                ║
║                                                              ║
║  GUIA              Cómo usar esta Excel (colores, flujo)     ║
║  Cap30_Marco       D1-D4, EUC, Tabla 44-47                   ║
║  Cap31_Teoria      MUI, I, ρₕ, EUC, T50 convergencia        ║
║  Cap31_Ejemplos    María, Tabla 49, ρₕ cálculo               ║
║  Cap32_Teoria      α, M, R, AHP T52-56, HITL T60            ║
║  Cap32_Ejemplos    T57-59 préstamo/mod/médico, T61 erosión   ║
║  Cap33_Teoria      CEE,ICD,IDP,ICT, SHAP T64, T65,T69       ║
║  Cap33_Ejemplos    ICT completo T74                           ║
║  Cap34_Teoria      Sigmoide T76, φ T77, pesos T81            ║
║  Cap34_Ejemplos    T78-83, T85 crédito, Tasa A (Ej24+25)    ║
║  Cap35_IHG         AHP T90-91, IHG, 3 condiciones, T92      ║
║  Cap35_Validacion  Patologías T93, sensibilidad T94, erosión ║
║  Dashboard         Panel resumen + dictamen certificación     ║
║                                                              ║
║  🟡 AMARILLO = editable   ⬜ GRIS = pesos autor              ║
║  🔵 AZUL = fórmula        🟢 VERDE = resultado               ║
║  🟣 LILA = referencia     🔴 ROJO = alerta                   ║
╚══════════════════════════════════════════════════════════════╝
""")
